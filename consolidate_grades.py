#!/usr/bin/env python3
"""
Combine Moodle rubric exports into a single consolidated workbook.

Workflow (per assignment file with numeric prefix other than 00):
- Strip rubric-only columns (Username, Score, Feedback, Graded by, Time graded).
- Drop any row containing "Raffi".
- Rename each remaining rubric column labelled "Definition" to a criterion name
  pulled from the matching rubric CSV (same filename minus numeric prefix) or
  fall back to Criterion 1..N.
- Add a total column that sums the criteria (labelled with the assignment weight).
- Pull the over-100 grade and letter grade from the 00-*.xlsx course total file.
- Arrange assignment blocks side-by-side with student names only once on the left,
  rotate headers 90°, and merge the assignment label above the criteria/totals.
- Append course total (over 100 and letter) to the far right and style with gray
  headers, alternating row shading, and a course title banner.
"""

from __future__ import annotations

import argparse
import re
import shutil
import subprocess
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BANNED_HEADERS = ["username", "score", "feedback", "graded by", "time graded"]
HEADER_TEXT_ROTATION = 90


def find_header_row(df: pd.DataFrame) -> int:
    """Locate the row index that contains the rubric header ("First name")."""
    matches = df[
        df.apply(
            lambda row: row.astype(str)
            .str.contains("First name", case=False, na=False)
            .any(),
            axis=1,
        )
    ]
    if matches.empty:
        raise ValueError("Could not locate the header row with 'First name'.")
    return matches.index[0]


def normalize_title(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip().lower()


def find_rubric_csv(path: Path) -> Optional[Path]:
    """Locate the rubric CSV that matches an assignment XLSX (without numeric prefix)."""
    name = path.name
    base_part = name.split("-", 1)[1] if "-" in name else path.stem
    base_part = base_part.rsplit(".", 1)[0].strip()
    direct = path.with_name(f"{base_part} - Rubric Percentage.csv")
    if direct.exists():
        return direct
    alt = path.with_name(f"{base_part}.csv")
    if alt.exists():
        return alt
    for candidate in path.parent.glob("*.csv"):
        if base_part in candidate.stem:
            return candidate
    return None


def load_criterion_labels(csv_path: Path, expected_count: int) -> List[str]:
    """Flatten rubric CSV cells (header + values) into ordered criterion labels."""
    try:
        df = pd.read_csv(csv_path)
    except Exception:
        return []
    labels: List[str] = []
    for col in df.columns:
        if len(labels) >= expected_count:
            break
        if pd.notna(col) and str(col).strip():
            labels.append(str(col).strip())
        for val in df[col]:
            if len(labels) >= expected_count:
                break
            if pd.isna(val):
                continue
            text = str(val).strip()
            if text:
                labels.append(text)
    return labels[:expected_count]


def extract_weight(title: str) -> Optional[str]:
    match = re.search(r"(\d+(?:\.\d+)?)%\s*$", title)
    if match:
        return f"{match.group(1)}%"
    match = re.search(r"(\d+(?:\.\d+)?)%", title)
    return f"{match.group(1)}%" if match else None


def strip_assignment_word(title: str) -> str:
    cleaned = re.sub(r"(?i)^assignment\s*[:-]?\s*", "", title).strip()
    return cleaned or title


def remove_raffi_rows(df: pd.DataFrame) -> pd.DataFrame:
    mask = df.apply(
        lambda row: row.astype(str).str.contains("Raffi", case=False, na=False).any(),
        axis=1,
    )
    return df[~mask]


def build_roster(course_df: pd.DataFrame, assignment_paths: Iterable[Path]) -> List[Tuple[str, str]]:
    roster: List[Tuple[str, str]] = []
    seen = set()

    def add_name(first: str, last: str) -> None:
        key = (first.strip(), last.strip())
        if key[0] and key[1] and key not in seen:
            roster.append(key)
            seen.add(key)

    for first, last in zip(course_df["First name"], course_df["Last name"]):
        add_name(str(first), str(last))

    for path in assignment_paths:
        raw = pd.read_excel(path, header=None, engine="openpyxl")
        header_row = find_header_row(raw)
        names = raw.iloc[header_row + 1 :, :2].dropna(how="all")
        names = remove_raffi_rows(names)
        for _, row in names.iterrows():
            add_name(str(row.iloc[0]), str(row.iloc[1]))

    return roster


def load_course_totals(course_path: Path) -> pd.DataFrame:
    course_df = pd.read_excel(course_path, engine="openpyxl")
    course_df = course_df[course_df["First name"].notna()]
    course_df[["First name", "Last name"]] = course_df[["First name", "Last name"]].apply(
        lambda col: col.astype(str).str.strip()
    )
    course_df = remove_raffi_rows(course_df)
    return course_df


def find_course_grade_columns(course_df: pd.DataFrame, assignment_title: str) -> Tuple[Optional[str], Optional[str]]:
    perc_col = None
    letter_col = None
    normalized_title = normalize_title(assignment_title)
    stripped_title = normalize_title(strip_assignment_word(assignment_title))

    for col in course_df.columns:
        col_norm = normalize_title(str(col))
        if "(percentage)" in col_norm and (normalized_title in col_norm or stripped_title in col_norm):
            perc_col = col
        if "(letter)" in col_norm and (normalized_title in col_norm or stripped_title in col_norm):
            letter_col = col
    return perc_col, letter_col


def parse_course_percentage(value):
    if pd.isna(value):
        return pd.NA
    if isinstance(value, str):
        value = value.replace("%", "").strip()
    try:
        return float(value)
    except Exception:
        return value


def align_to_roster(df: pd.DataFrame, roster: List[Tuple[str, str]]) -> pd.DataFrame:
    idx = pd.MultiIndex.from_tuples(roster, names=["First name", "Last name"])
    aligned = df.set_index(["First name", "Last name"]).reindex(idx)
    return aligned.reset_index()


def parse_assignment(path: Path, course_df: pd.DataFrame, roster: List[Tuple[str, str]]):
    raw = pd.read_excel(path, header=None, engine="openpyxl")
    assignment_title = str(raw.iloc[1, 0]).strip()
    weight = extract_weight(assignment_title)
    header_row = find_header_row(raw)
    header = list(raw.iloc[header_row])
    data = raw.iloc[header_row + 1 :].dropna(how="all").reset_index(drop=True)
    data.columns = header

    data = remove_raffi_rows(data)

    keep_indices = [
        idx
        for idx, col in enumerate(header)
        if not (isinstance(col, str) and any(bad in col.lower() for bad in BANNED_HEADERS))
    ]
    cleaned = data.iloc[:, keep_indices].copy()

    columns: List[str] = []
    criterion_cols: List[str] = []
    definition_count = sum(
        1 for label in cleaned.columns if isinstance(label, str) and label.lower() == "definition"
    )
    csv_path = find_rubric_csv(path)
    criterion_labels = load_criterion_labels(csv_path, definition_count) if csv_path else []
    criterion_idx = 0
    for label in cleaned.columns:
        if isinstance(label, str) and label.lower() == "definition":
            if criterion_idx < len(criterion_labels):
                new_name = criterion_labels[criterion_idx]
            else:
                new_name = f"Criterion {len(criterion_cols) + 1}"
            criterion_idx += 1
            criterion_cols.append(new_name)
            columns.append(new_name)
        elif isinstance(label, str) and label.lower() == "first name":
            columns.append("First name")
        elif isinstance(label, str) and label.lower() == "last name":
            columns.append("Last name")
        else:
            columns.append(str(label) if not pd.isna(label) else "")
    cleaned.columns = columns

    cleaned["First name"] = cleaned["First name"].astype(str).str.strip()
    cleaned["Last name"] = cleaned["Last name"].astype(str).str.strip()

    for col in criterion_cols:
        cleaned[col] = pd.to_numeric(cleaned[col], errors="coerce")

    total_label = f"Total - {weight}" if weight else "Total"
    cleaned[total_label] = cleaned[criterion_cols].sum(axis=1, numeric_only=True)

    perc_col, letter_col = find_course_grade_columns(course_df, assignment_title)
    course_indexed = course_df.set_index(["First name", "Last name"])

    def lookup(name_tuple: Tuple[str, str], column: Optional[str]):
        if column is None or name_tuple not in course_indexed.index:
            return pd.NA
        return course_indexed.at[name_tuple, column]

    cleaned["Total - 100"] = [
        parse_course_percentage(lookup((fn, ln), perc_col)) for fn, ln in zip(cleaned["First name"], cleaned["Last name"])
    ]
    cleaned["Total - Letter"] = [
        lookup((fn, ln), letter_col) for fn, ln in zip(cleaned["First name"], cleaned["Last name"])
    ]

    aligned = align_to_roster(cleaned, roster)

    columns_order = ["First name", "Last name"] + criterion_cols + [total_label, "Total - 100", "Total - Letter"]
    aligned = aligned[columns_order]
    write_columns = criterion_cols + [total_label, "Total - 100", "Total - Letter"]

    return {
        "title": assignment_title,
        "display_name": strip_assignment_word(assignment_title),
        "weight": weight,
        "columns": columns_order,
        "write_columns": write_columns,
        "df": aligned,
    }


def course_total_columns(course_df: pd.DataFrame) -> Tuple[Optional[str], Optional[str]]:
    perc_col = None
    letter_col = None
    for col in course_df.columns:
        norm = normalize_title(str(col))
        if "course total" in norm and "(percentage)" in norm:
            perc_col = col
        if "course total" in norm and "(letter)" in norm:
            letter_col = col
    return perc_col, letter_col


def course_totals_for_roster(course_df: pd.DataFrame, roster: List[Tuple[str, str]]):
    perc_col, letter_col = course_total_columns(course_df)
    indexed = course_df.set_index(["First name", "Last name"])
    perc_values = []
    letter_values = []
    for fn, ln in roster:
        key = (fn, ln)
        if perc_col and key in indexed.index:
            perc_values.append(parse_course_percentage(indexed.at[key, perc_col]))
        else:
            perc_values.append(pd.NA)
        if letter_col and key in indexed.index:
            letter_values.append(indexed.at[key, letter_col])
        else:
            letter_values.append(pd.NA)
    return perc_values, letter_values


def read_course_name(sample_rubric_path: Path) -> str:
    raw = pd.read_excel(sample_rubric_path, header=None, engine="openpyxl")
    value = raw.iloc[0, 0]
    return str(value).strip() if pd.notna(value) else "Course"


def write_workbook(assignments: List[dict], roster: List[Tuple[str, str]], course_df: pd.DataFrame, course_name: str, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated"

    total_columns_count = 2 + sum(len(a["write_columns"]) for a in assignments) + 2

    title_row = 1
    group_row = 2
    header_row = 3  # rotated headers
    data_start = header_row + 1
    col_offset = 3  # start after name columns

    header_alignment = Alignment(text_rotation=HEADER_TEXT_ROTATION, horizontal="center", vertical="bottom", wrap_text=True)
    group_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    name_alignment = Alignment(horizontal="left", vertical="center")

    title_fill = PatternFill("solid", fgColor="BFBFBF")  # dark gray
    header_fill = PatternFill("solid", fgColor="E6E6E6")  # light gray
    group_fill = PatternFill("solid", fgColor="D0D0D0")  # mid gray
    stripe_fill = PatternFill("solid", fgColor="F7F7F7")  # zebra alternate
    thin = Side(style="thin", color="999999")
    thick = Side(style="medium", color="666666")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    def clean_value(value):
        return None if pd.isna(value) else value

    def set_cell(cell, value, align=None, fill=None, bold=False, number=False, border_on=True):
        cell.value = clean_value(value)
        if align:
            cell.alignment = align
        if fill:
            cell.fill = fill
        if bold:
            cell.font = Font(bold=True, size=cell.font.size if cell.font else 11)
        if number and isinstance(value, (int, float)):
            cell.number_format = "0.00"
        if border_on:
            cell.border = border

    # Top banner with course name and subtitle.
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=total_columns_count)
    top = ws.cell(row=title_row, column=1, value=f"{course_name}\nAssignment Grade Breakdown Per Criteria")
    top.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top.font = Font(bold=True, size=13)
    top.fill = title_fill
    top.border = border

    # Names block
    ws.merge_cells(start_row=group_row, start_column=1, end_row=group_row, end_column=2)
    set_cell(ws.cell(row=group_row, column=1), "Students", align=group_alignment, fill=group_fill, bold=True)
    set_cell(ws.cell(row=header_row, column=1), "First name", align=header_alignment, fill=header_fill, bold=True)
    set_cell(ws.cell(row=header_row, column=2), "Last name", align=header_alignment, fill=header_fill, bold=True)

    for idx, (first, last) in enumerate(roster):
        row = data_start + idx
        set_cell(ws.cell(row=row, column=1), first, align=name_alignment)
        set_cell(ws.cell(row=row, column=2), last, align=name_alignment)

    section_ends = [2]  # end of names block

    for assignment in assignments:
        df = assignment["df"]
        columns = assignment["write_columns"]

        # Write column headers with rotation
        for idx, col_name in enumerate(columns):
            set_cell(ws.cell(row=header_row, column=col_offset + idx), col_name, align=header_alignment, fill=header_fill, bold=True)

        # Merge and label the assignment group above the criteria/totals.
        merge_start = col_offset  # start at Criterion 1
        merge_end = col_offset + len(columns) - 1  # through Total - Letter
        if merge_start <= merge_end:
            ws.merge_cells(start_row=group_row, start_column=merge_start, end_row=group_row, end_column=merge_end)
            set_cell(
                ws.cell(row=group_row, column=merge_start),
                assignment["display_name"],
                align=group_alignment,
                fill=group_fill,
                bold=True,
            )

        # Dump student rows.
        for row_idx, (_, row) in enumerate(df.iterrows()):
            for col_idx, col_name in enumerate(columns):
                set_cell(
                    ws.cell(row=data_start + row_idx, column=col_offset + col_idx),
                    row[col_name],
                    align=data_alignment,
                    number=True,
                )

        section_ends.append(col_offset + len(columns) - 1)
        col_offset += len(columns)

    # Course total block
    course_perc, course_letter = course_totals_for_roster(course_df, roster)
    total_columns = ["Course total - 100", "Course total - Letter"]
    ws.merge_cells(start_row=group_row, start_column=col_offset, end_row=group_row, end_column=col_offset + 1)
    set_cell(ws.cell(row=group_row, column=col_offset), "Course Total", align=group_alignment, fill=group_fill, bold=True)

    for idx, col_name in enumerate(total_columns):
        set_cell(ws.cell(row=header_row, column=col_offset + idx), col_name, align=header_alignment, fill=header_fill, bold=True)
    for row_idx, (perc, letter) in enumerate(zip(course_perc, course_letter)):
        row_num = data_start + row_idx
        for offset, value in enumerate((perc, letter)):
            set_cell(ws.cell(row=row_num, column=col_offset + offset), value, align=data_alignment, number=True)
    section_ends.append(col_offset + 1)

    # Zebra striping for data rows.
    max_col = ws.max_column
    for row_idx in range(len(roster)):
        if row_idx % 2 == 1:
            row_num = data_start + row_idx
            for col in range(1, max_col + 1):
                ws.cell(row=row_num, column=col).fill = stripe_fill

    # Styling tweaks
    ws.freeze_panes = ws["C4"]

    # Autosize columns to content (within bounds).
    def autosize():
        header_max_len = 0
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=1, max_row=ws.max_row):
                for c in cell:
                    if c.value is None:
                        continue
                    val = str(c.value).replace("\n", " ")
                    max_len = max(max_len, len(val))
                    if c.row in (title_row, group_row, header_row):
                        header_max_len = max(header_max_len, len(val))
            # Names columns get a wider default cap.
            if col <= 2:
                min_w, max_w = 14, 40
            else:
                min_w, max_w = 8, 30
            width = max(min_w, max_len + 2)
            width = min(width, max_w)
            ws.column_dimensions[get_column_letter(col)].width = width

        # Adjust header row height based on rotated text length to avoid clipping.
        ws.row_dimensions[header_row].height = max(60, header_max_len * 2)
        ws.row_dimensions[group_row].height = max(ws.row_dimensions[group_row].height or 0, 24)

    # Add thick vertical separators between sections.
    def add_thick_vertical(col_idx: int):
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=cell.border.left,
                right=thick,
                top=cell.border.top,
                bottom=cell.border.bottom,
            )

    for end_col in section_ends:
        add_thick_vertical(end_col)

    autosize()

    # Page setup for PDF export: single page, landscape.
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_options.horizontalCentered = True

    wb.save(output_path)
    return wb


def export_pdf(xlsx_path: Path, pdf_path: Path) -> bool:
    """Export the XLSX to PDF via LibreOffice/soffice if available."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        print(f"Skipping PDF export (LibreOffice/soffice not found). Intended path: {pdf_path}")
        return False

    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        subprocess.run(
            [
                soffice,
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(pdf_path.parent),
                str(xlsx_path),
            ],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        # LibreOffice names the output based on the input filename.
        produced = xlsx_path.with_suffix(".pdf")
        if produced != pdf_path:
            if produced.exists():
                produced.replace(pdf_path)
        return True
    except subprocess.CalledProcessError:
        print(f"PDF export failed via LibreOffice for {xlsx_path}")
        return False


def list_assignment_files(base_dir: Path) -> Tuple[Path, List[Path]]:
    course_file = None
    rubric_files = []
    for path in base_dir.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        prefix = path.name.split("-", 1)[0]
        if prefix == "00":
            course_file = path
        elif prefix.isdigit():
            rubric_files.append(path)
    rubric_files = [p for p in rubric_files if p.name.split("-", 1)[0] != "00"]
    rubric_files.sort(key=lambda p: int(p.name.split("-", 1)[0]))
    if course_file is None:
        raise FileNotFoundError("No course total file starting with '00-' was found.")
    return course_file, rubric_files


def main() -> None:
    parser = argparse.ArgumentParser(description="Unify Moodle rubric exports into a consolidated gradebook.")
    parser.add_argument("--input-dir", default=".", help="Directory containing the XLSX exports (default: current).")
    parser.add_argument(
        "--output", default="consolidated.xlsx", help="Path for the consolidated workbook (default: consolidated.xlsx)."
    )
    args = parser.parse_args()

    base_dir = Path(args.input_dir).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    course_path, rubric_paths = list_assignment_files(base_dir)
    course_df = load_course_totals(course_path)
    roster = build_roster(course_df, rubric_paths)
    course_name = read_course_name(rubric_paths[0]) if rubric_paths else "Course"

    assignments = []
    for path in rubric_paths:
        assignment = parse_assignment(path, course_df, roster)
        assignments.append(assignment)

    write_workbook(assignments, roster, course_df, course_name, output_path)
    print(f"Wrote consolidated workbook to {output_path}")

    pdf_path = output_path.with_suffix(".pdf")
    if export_pdf(output_path, pdf_path):
        print(f"Wrote PDF to {pdf_path}")
    else:
        print("PDF not produced.")


if __name__ == "__main__":
    main()
